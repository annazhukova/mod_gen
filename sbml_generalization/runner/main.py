#!/usr/bin/env python
# encoding: utf-8

import glob
import logging
import os

import libsbml
import openpyxl
import openpyxl.styles
from openpyxl.styles import Style, Font

from mod_sbml.annotation.chebi.chebi_annotator import get_chebi_term_by_annotation
from mod_sbml.annotation.chebi.chebi_serializer import get_chebi
from mod_sbml.onto import parse_simple
from mod_sbml.sbml.sbml_manager import get_gene_association
from mod_sbml.serialization import get_sbml_r_formula
from mod_sbml.utils.misc import invert_map
from sbml_generalization.generalization.sbml_generalizer import generalize_model
from sbml_generalization.merge.model_merger import merge_models

__author__ = 'anna'

BASIC_STYLE = Style(font=Font(color=openpyxl.styles.colors.BLACK, sz=8))
BOLD_STYLE = Style(font=Font(bold=True, sz=8))
HEADER_STYLE = BOLD_STYLE


def add_values(ws, row, col, values, style=BASIC_STYLE):
    j = col
    for v in values:
        ws.cell(row=row, column=j).value = v
        ws.cell(row=row, column=j).style = style
        j += 1


def serialize_generalization(r_id2clu, s_id2clu, sbml, chebi, path):
    doc = libsbml.SBMLReader().readSBML(sbml)
    model = doc.getModel()
    groups_plugin = model.getPlugin("groups")
    clu2r_ids, clu2s_ids = invert_map(r_id2clu), invert_map(s_id2clu)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(0, "Metabolite Groups")
    row = 1
    add_values(ws, row, 1, ["Group Id", "Group Name", "Group CHEBI", "Id", "Name", "Compartment", "CHEBI"], HEADER_STYLE)
    row += 1
    processed_s_ids = set()
    for (g_id, ch_term), s_ids in sorted(clu2s_ids.items(), key=lambda ((g_id, _), s_ids): g_id):
        group = groups_plugin.getGroup(g_id)
        add_values(ws, row, 1, [g_id, group.getName(), ch_term.get_id() if ch_term else ''])
        for s_id in sorted(s_ids, key=lambda s_id: s_id[s_id.find('__'):]):
            species = model.getSpecies(s_id)
            ch_term = get_chebi_term_by_annotation(species, chebi)
            add_values(ws, row, 4, [s_id, species.getName(), model.getCompartment(species.getCompartment()).getName(),
                                    ch_term.get_id() if ch_term else ''])
            row += 1
        processed_s_ids |= s_ids
    ws = wb.create_sheet(1, "Ungrouped metabolites")
    row = 1
    add_values(ws, row, 1, ["Id", "Name", "Compartment", "CHEBI"], HEADER_STYLE)
    row += 1
    unm_l = 0
    for species in sorted(model.getListOfSpecies(), key=lambda s: s.getId()[s.getId().find('__'):]):
        if species.getId() not in processed_s_ids:
            ch_term = get_chebi_term_by_annotation(species, chebi)
            add_values(ws, row, 1, [species.getId(), species.getName(),
                                    model.getCompartment(species.getCompartment()).getName(),
                                    ch_term.get_id() if ch_term else ''])
            row += 1
            unm_l += 1
    print unm_l

    ws = wb.create_sheet(2, "Reaction Groups")
    row = 1
    add_values(ws, row, 1, ["Group Id", "Id", "Name", "Formula", "Gene Association"], HEADER_STYLE)
    row += 1
    processed_r_ids = set()
    for (g_id, g_name), r_ids in sorted(clu2r_ids.items(), key=lambda ((g_id, g_name), _): g_id):
        add_values(ws, row, 1, [g_id])
        for r_id in sorted(r_ids, key=lambda r_id: r_id[r_id.find('__'):]):
            r = model.getReaction(r_id)
            add_values(ws, row, 2, [r_id, r.getName(), get_sbml_r_formula(model, r, show_compartments=True, show_metabolite_ids=True),
                                    get_gene_association(r)])
            row += 1
        processed_r_ids |= r_ids
    ws = wb.create_sheet(3, "Ungrouped reactions")
    row = 1
    add_values(ws, row, 1, ["Id", "Name", "Formula", "Gene Association"], HEADER_STYLE)
    row += 1
    unm_l = 0
    for r in sorted(model.getListOfReactions(), key=lambda r: r.getId()[r.getId().find('__'):]):
        if r.getId() not in processed_r_ids:
            add_values(ws, row, 1, [r.getId(), r.getName(), get_sbml_r_formula(model, r, show_compartments=True, show_metabolite_ids=True),
                                    get_gene_association(r)])
            row += 1
            unm_l += 1
    print unm_l

    wb.save(path)


if __name__ == "__main__":

    # parameter parsing #
    import argparse

    parser = argparse.ArgumentParser(description="Generalizes an SBML model.")
    parser.add_argument('--model', required=True, type=str,
                        help="input model in SBML format "
                             "or a directory containing several models if they first need to be merged")
    parser.add_argument('--chebi', default=None, type=str, help="path to the ChEBI ontology file in OBO format")
    parser.add_argument('--output_model', default=None, type=str,
                        help="path to the output generalized model in SBML format")
    parser.add_argument('--groups_model', default=None, type=str,
                        help="path to the output model in SBML format with groups extension to encode similar elements")
    parser.add_argument('--merged_model', default=None, type=str,
                        help="path to the output merged model in SBML format")
    parser.add_argument('--verbose', action="store_true", help="print logging information")
    parser.add_argument('--log', default=None, help="a log file")
    params = parser.parse_args()

    if not params.chebi:
        params.chebi = get_chebi()
    prefix = os.path.splitext(params.model)[0]
    if not params.merged_model:
        params.merged_model = "%s.xml" % prefix
    if not params.output_model:
        params.output_model = "%s_generalized.xml" % prefix
    if not params.groups_model:
        params.groups_model = "%s_with_groups.xml" % prefix

    if params.verbose:
        logging.basicConfig(level=logging.INFO)

    logging.info("parsing ChEBI...")
    ontology = parse_simple(params.chebi)
    if os.path.isdir(params.model):
        in_sbml_list = ['%s/%s' % (params.model, f) for f in glob.glob(os.path.join(params.model, '*'))
                        if os.path.splitext(f)[1] in [".xml", ".sbml"]]
        for sbml in in_sbml_list:
            doc = libsbml.SBMLReader().readSBML(sbml)
            model = doc.getModel()
            print(sbml, model.getNumSpecies(), model.getNumReactions())
        merge_models(in_sbml_list, params.merged_model)
        sbml = params.merged_model
    else:
        sbml = params.model
    r_id2clu, s_id2clu, _, _ = generalize_model(sbml, ontology, params.groups_model, params.output_model,
                                                ub_chebi_ids={'chebi:ch'})
    if os.path.isdir(params.model):
        serialize_generalization(r_id2clu, s_id2clu, params.groups_model, ontology,
                                 os.path.join(params.model, 'generalized.xlsx'))

